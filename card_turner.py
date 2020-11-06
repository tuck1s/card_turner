#!/usr/bin/env python3
# Simple card turning app
import argparse, configparser, random, time
from tkinter import *
from datetime import datetime
import openpyxl
from openpyxl.styles import colors
from theme_and_tint_to_rgb import theme_and_tint_to_rgb

class Deck():
    """A deck comprises cards that can be shuffled, drawn, undrawn. Drawn cards go onto the discard pile"""

    def __init__(self, name, color='white', textcolor='black'):
        self.name = name
        self.cards = []
        self.discards = []
        self.table = None
        self.color = color
        self.textcolor = textcolor
        return

    def add(self, c):
        """Add a single card, or a list of cards"""
        if iter(c) and not isinstance(c, str):
            self.cards.extend(c)
        else:
            self.cards.append(c)

    def shuffle(self):
        """Randomise the order of the cards and the discards"""
        random.shuffle(self.cards)
        random.shuffle(self.discards)
        return

    def draw(self):
        """Move a card from the front of the deck to the table. If exhausted, returns None """
        try:
            self.discard() # ensure the table is empty
            self.table = self.cards.pop()
        finally:
            return self.label(self.table)

    def discard(self):
        """Move card from table to discard. If None, does nothing """
        if self.table:
            self.discards.append(self.table)
            self.table = None
            return self.label(self.table)

    def undiscard(self):
        """ Get a card back from the discards to the table. If exhausted, returns None """
        try:
            self.undraw() # ensure the table is empty
            self.table = self.discards.pop()
        finally:
            return self.label(self.table)


    def undraw(self):
        """ Get a card back from the discards to the table. If exhausted, returns empty """
        if self.table:
            self.cards.append(self.table)
            self.table = None
        return self.label(self.table)

    def label(self, c):
        if c == None:
            return ''
        else:
            return c

# -----------------------------------------------------------------------------------------

class App():
    def handle_draw(self, event):
        num = event.widget.extra[0]
        lbl = event.widget.extra[1]
        c = self.decks[num].draw()
        lbl['text'] = c
        self.reset_clock()


    def handle_undraw(self, event):
        num = event.widget.extra[0]
        lbl = event.widget.extra[1]
        c = self.decks[num].undraw()
        lbl['text'] = c
        self.reset_clock()


    def handle_discard(self, event):
        num = event.widget.extra[0]
        lbl = event.widget.extra[1]
        c = self.decks[num].discard()
        lbl['text'] = c
        self.reset_clock()


    def handle_undiscard(self, event):
        num = event.widget.extra[0]
        lbl = event.widget.extra[1]
        c = self.decks[num].undiscard()
        lbl['text'] = c
        self.reset_clock()


    def reset_clock(self):
        self.time_started = datetime.now()
        self.timer_label['text'] = '00:00'
        self.window.after(1000, self.update_clock)

    def update_clock(self):
        spent = datetime.now() - self.time_started
        minutes, seconds = divmod(spent.seconds, 60)
        self.timer_label['text'] = '{:02}:{:02}'.format(int(minutes), int(seconds))
        self.window.after(1000, self.update_clock)

    def __init__(self, title, decks, bg):
        self.window = Tk()
        self.window.attributes('-topmost', 1) # bring window in front of others
        self.window.title(title)
        self.decks = decks

        f_width = 100
        frm = Frame(master=self.window, width=f_width, bg=bg)
        self.timer_label = Label(master=frm)
        self.reset_clock()

        self.timer_label.pack()

        for i, d in enumerate(decks):
            frm.grid(row=i, column=1, padx=5, pady=5)

            # f1: Deck names
            f1 = Frame(master=frm, width=f_width, relief=GROOVE, borderwidth=5, bg=bg)
            lbl1 = Label(master=f1, text=d.name, bg=d.color, fg=d.textcolor, width=30, wraplength=f_width)
            lbl1.pack(fill=X)

            # f2: Draw / undraw buttons
            f2 = Frame(master=frm, width=f_width, bg=bg)
            btn1 = Button(master=f2, text='⬇︎', highlightbackground=bg)
            btn1.bind('<Button-1>', self.handle_draw)
            btn1.pack(side=LEFT)

            btn2 = Button(master=f2, text='⬆︎', highlightbackground=bg)
            btn2.bind('<Button-1>', self.handle_undraw)
            btn2.pack(side=RIGHT)

            # f3: Table, has one card face-up
            f3 = Frame(master=frm, width=f_width, bg=bg)
            lbl2 = Label(master=f3, text=None, bg=d.color, fg=d.textcolor, wraplength=200)
            lbl2.pack()

            # f4: Discard / undiscard buttons
            f4 = Frame(master=frm, width=f_width, bg=bg)
            btn3 = Button(master=f4, text='⬇︎', highlightbackground=bg)
            btn3.bind('<Button-1>', self.handle_discard)
            btn3.pack(side=LEFT)

            btn4 = Button(master=f4, text='⬆︎', bg='gray', highlightbackground=bg)
            btn4.bind('<Button-1>', self.handle_undiscard)
            btn4.pack(side=RIGHT)

            f1.pack(side='top', fill='both', expand=True)
            f2.pack(fill='both', expand=True)
            f3.pack(fill='both', expand=True)
            f4.pack(fill='both', expand=True)

            # Provide hooks to deck number and label
            btn1.extra = (i, lbl2)
            btn2.extra = (i, lbl2)
            btn3.extra = (i, lbl2)
            btn4.extra = (i, lbl2)
        self.window.mainloop()

# -----------------------------------------------------------------------------------------

def get_decks(f):
    """ Get card deck categories, colors and text contents from Excel file. Always loads the first sheet i.e. the active one """

    wb = openpyxl.load_workbook(f.name)
    ws = wb.active
    # Pick up deck names from the first row
    decks = []
    for j in range(1, ws.max_column+1):
        i = ws.cell(column=j, row=1)
        name = i.value
        if name:
            text_color = get_color(wb, i.font.color)
            bg_color = get_color(wb, i.fill.start_color)
            d = Deck(name, bg_color, text_color)

            # Pick up the cards from each column
            cards = []
            area = ws.iter_cols(min_row=2, min_col=j, max_col=j, values_only=True)
            for col in area:
                for item in col:
                    if item != None:
                        d.add(stripquotes(item))

            print('{:20} {} cards'.format(d.name, len(d.cards)))
            d.shuffle()
        decks.append(d)
    return  ws.title, decks

def get_color(wb, c):
    """ Returns the RGB color from an Excel Color attribute, stripping off the alpha """
    if c.type == 'rgb':
        argb = c.rgb
    elif c.type == 'theme':
        argb = theme_and_tint_to_rgb(wb, c.theme, c.tint)
    elif c.type == 'indexed':
        argb = colors.COLOR_INDEX[c.value] # best effort, untested
    else:
        raise Exception('Unknown Excel color type')
    # map 'no fill' to white
    return 'white' if argb == '00000000' else '#' + argb[-6:]

def stripquotes(s):
    """ Enforce to be a string and strip matching quotes """
    s = str(s)
    return s.lstrip('"').rstrip('"')

# -----------------------------------------------------------------------------------------
# Main code
# -----------------------------------------------------------------------------------------
parser = argparse.ArgumentParser(description='Simple card turning game app')
parser.add_argument('excelfile', metavar='file.xlsx', type=argparse.FileType('r'),
                    help='Excel input file')
parser.add_argument('-bg', '--background', type=str, action='store', default='#000000',
                    help='Background color e.g. black, 202020')
args = parser.parse_args()
title, decks = get_decks(args.excelfile)
print('{} card decks read'.format(len(decks)))
App(title, decks, args.background)

